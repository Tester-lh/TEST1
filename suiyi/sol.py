# -*- coding: utf-8 -*-
__author__ = 'NL-lh'

import yaml
from openpyxl import load_workbook

wb = load_workbook(filename='sol.xlsx')
sheets = wb.get_sheet_names()  # 获取所有表格(worksheet)的名字
sheet_0 = sheets[0]  # 第一个表格的名称
ws = wb.get_sheet_by_name(sheet_0)  # 获取特定的 worksheet

# 获取表格所有行和列，两者都是可迭代的
rows = ws.rows
columns = ws.columns
i = 1
s_id = []
for row in rows:
    s_id.append(ws.cell(row=i, column=1).value)
    i = i + 1
temp = yaml.load(open("data.yml", 'r'))
print(temp)
for s in range(len(s_id)):
    print(temp[s_id[s]])
    ws.cell(row=s+1, column=5).value=temp[s_id[s]]

wb.save(filename='sol.xlsx')
