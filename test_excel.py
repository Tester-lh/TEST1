from openpyxl import  Workbook
wb=Workbook()
ws=wb.get_active_sheet()
print(ws.title)
ws.title='My Sheet1'

ws['D6'].value=4
ws.cell(row=3, column=1).value=6

new_ws=wb.create_sheet(title='my sheet2')
for row in range(1,101):
    for col in range(1,11):
        new_ws.cell(row=row,column=col).value=row+col

new_ws2=wb.create_sheet(title='homework99')
for i in range(1,10):
    for j in range(1,10):
        if j<=i:
            new_ws2.cell(row=i+2,column=j+3).value=i*j


wb.save(filename='testExcel.xlsx')
# wb.save('test.xlsx')也行

