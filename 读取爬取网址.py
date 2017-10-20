from openpyxl import load_workbook
from selenium import webdriver
import urllib
import urllib.request
import http.cookiejar

myExcel = load_workbook(filename='业务大全网址爬取.xlsx')
sheets = myExcel.get_sheet_names()  # 获取所有表格(worksheet)的名字
sheet_0 = sheets[0]  # 第一个表格的名称
mySheet = myExcel.get_sheet_by_name(sheet_0)  # 获取特定的 worksheet
rows = mySheet.rows
columns = mySheet.columns
i = 2
link = []
for row in rows:
    link.append(str(mySheet.cell(row=i, column=2).value))
    print(link[i - 2])
    i += 1

browser = webdriver.Firefox()
cookie = browser.get_cookies()
for links in link:
    if links is not None:
        url = links
        print(url)
        browser.get(url)
        browser.delete_all_cookies()
        cookie_jar = http.cookiejar.MozillaCookieJar()
        cookies = open('cookies.txt').read()
        browser.manager().addCookie(cookies)
        print(browser.get_cookies())
    break

    # cookie_jar中已经加载了当前页面的Cookie数据了，Enjoy!
    # # get the session cookie
    # cookie = [item["name"] + "=" + item["value"] for item in sel.get_cookies()]
    # # print cookie
    #
    # cookiestr = ';'.join(item for item in cookie)
    # print(cookiestr)
    #
    #
    # homeurl = sel.current_url
    # print('homeurl: %s' % homeurl)
    # headers = {'cookie':cookiestr}
    # req = urllib.request.Request(homeurl, headers = headers)
    # try:
    #     response = urllib.request.urlopen(req)
    #     text = response.read()
    #     fd = open('homepage', 'w')
    #     fd.write(text)
    #     fd.close()
    #     print('###get home page html success!!')
    # except:
    #     print( '### get home page html error!!' )
