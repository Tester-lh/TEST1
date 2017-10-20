# encoding: utf-8

from selenium import webdriver
from openpyxl import Workbook

import urllib
wb = Workbook()
ws = wb.get_active_sheet()
ws.title = '业务大全网址爬取'
ws.cell(row=1, column=1).value = '业务名称'
ws.cell(row=1, column=2).value = '链接网址'


browser = webdriver.Chrome()
browser.set_page_load_timeout(10)
i = 2
url = 'http://www.sh.10086.cn/sh/wsyyt/busi/2010.jsp'
browser.get(url)
items = browser.find_elements_by_css_selector('#wrap > div > div.wrap-right > div.wrap-rig-con1 > div.box1 > div > div.class-box-content > div > div.class-box-right > div > ul > li > a')
for item in items:
    title = item.text
    link = item.get_attribute('href')
    print(title + ': '+link)
    ws.cell(row=i, column=1).value = title
    ws.cell(row=i, column=2).value = str(link)
    i+=1

wb.save(filename='业务大全网址爬取.xlsx')
