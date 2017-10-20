# encoding: utf-8

from openpyxl import load_workbook

wb = load_workbook(filename='testExcel.xlsx')

sheets = wb.get_sheet_names()  # 获取所有表格(worksheet)的名字
sheet_0 = sheets[0]  # 第一个表格的名称
ws = wb.get_sheet_by_name(sheet_0)  # 获取特定的 worksheet

# 获取表格所有行和列，两者都是可迭代的
rows = ws.rows
columns = ws.columns

# 行迭代
for row in rows:
    line = [col.value for col in row]
    print(line)

# 通过坐标读取值
# print(ws.cell('D5').value)  # D 表示列，5 表示行
print(ws['D5'].value)
print(ws.cell(row=5, column=4).value)